"""Schema (structured-data) audit + recommendation engine — the app-integrated layer.

Reuses the pure analysis library in `schema_checker/` (schema.org vocabulary validation,
Google rich-result rules, intent detection, recommendations, JSON-LD generation, scoring)
and the existing crawl plumbing (extractor, ingest sitemap discovery, settings throttling).

Adds what the spec requires: STRICT project scoping (every URL must belong to the active
source's domain), the two-mode ingestion (bulk import with optional current-schema diff;
sitemap auto-detect with manual selection), robots/throttle-respecting fetch, and storage
in `page_schema`. Nothing here audits until the caller passes an explicit URL selection.
"""
from __future__ import annotations

import json
import time
from datetime import datetime, timezone
from urllib.parse import urljoin, urlparse

import httpx

from ..config import settings
from ..db import now_iso, source_domain
from ..extractor import extract as page_extract
from ..extractor import structured_data
from ..schema_checker import generate, google_rules, intent, recommend, score, validate_vocab, vocab
from ..util import host_excluded, host_of, parse_locale_section, registrable_domain


# ── project scoping (critical) ─────────────────────────────────────────────
def scope_domain(conn, source_id: int) -> str | None:
    """The active project's registrable domain (sources.domain)."""
    return source_domain(conn, source_id)


def in_scope(url: str, domain: str, exclude_hosts=None) -> tuple[bool, str]:
    """In-scope = same registrable domain as the project AND not an excluded host.
    Accepts the domain, its subdomains, and its subfolders; rejects everything else."""
    if not url or not url.lower().startswith("http"):
        return False, "not an http(s) URL"
    if not host_of(url):
        return False, "no host"
    if registrable_domain(url) != domain:
        return False, f"different domain ({registrable_domain(url)} ≠ {domain}) — needs its own project"
    if host_excluded(url, exclude_hosts or settings()["crawl"].get("exclude_hosts") or []):
        return False, "host is excluded (settings.exclude_hosts)"
    return True, ""


def scope_filter(urls, domain, exclude_hosts=None):
    """Split a URL list into (in_scope, rejected[(url, reason)]) — deduped, order-preserving."""
    keep, rejected, seen = [], [], set()
    for u in urls:
        u = (u or "").strip()
        if not u or u in seen:
            continue
        seen.add(u)
        ok, reason = in_scope(u, domain, exclude_hosts)
        (keep if ok else rejected).append(u if ok else (u, reason))
    return keep, rejected


# ── bulk-import reader (paste + txt/csv/xlsx), optional current-schema column ──
def read_rows_from_text(text: str) -> list[tuple[str, str | None]]:
    """Paste mode: one URL per line; an optional current-schema after a tab/comma."""
    rows = []
    for line in (text or "").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or not line.lower().startswith("http"):
            continue
        # allow "url<TAB>{json}" or "url,{json}" (json may contain commas → split once on tab first)
        if "\t" in line:
            url, cur = line.split("\t", 1)
        else:
            url, cur = line, ""
        rows.append((url.strip(), cur.strip() or None))
    return rows


def read_rows_from_file(path: str) -> list[tuple[str, str | None]]:
    """Upload mode: .txt/.csv (col0=url, optional col1=current schema) or .xlsx (same)."""
    from pathlib import Path
    p = Path(path)
    ext = p.suffix.lower()
    rows: list[tuple[str, str | None]] = []
    if ext == ".xlsx":
        from openpyxl import load_workbook
        ws = load_workbook(p, read_only=True, data_only=True).active
        for r in ws.iter_rows(values_only=True):
            if not r:
                continue
            url = str(r[0]).strip() if r[0] else ""
            if url.lower().startswith("http"):
                cur = str(r[1]).strip() if len(r) > 1 and r[1] else None
                rows.append((url, cur))
    elif ext == ".csv":
        import csv
        with open(p, newline="", encoding="utf-8") as fh:
            for r in csv.reader(fh):
                if r and r[0].strip().lower().startswith("http"):
                    rows.append((r[0].strip(), (r[1].strip() if len(r) > 1 and r[1].strip() else None)))
    else:  # .txt and anything else: reuse the existing plain-list reader
        from ..ingest import _read_url_list
        rows = [(u, None) for u in _read_url_list(str(p))]
    return rows


# ── sitemap discovery (reuses ingest), scoped ─────────────────────────────
def discover_urls(conn, source_id: int, address: str) -> dict:
    """Auto-detect the project's sitemap and return IN-SCOPE URLs (grouped by section) +
    the rejected out-of-scope list. Does NOT fetch/audit anything (manual selection first)."""
    from ..ingest import _client, _collect_sitemap_entries, _discover_sitemaps, canonical_base
    domain = scope_domain(conn, source_id)
    # the entered address itself must be in scope
    ok, reason = in_scope(address if address.startswith("http") else f"https://{address}", domain)
    if not ok:
        return {"error": f"'{address}' is out of scope for this project ({domain}). {reason}",
                "domain": domain, "in_scope": [], "rejected": [], "sections": []}
    exclude = settings()["crawl"].get("exclude_hosts") or []
    with _client() as client:
        base = canonical_base(client, address if address.startswith("http") else f"https://{address}")
        sitemaps = _discover_sitemaps(client, base)
        entries = _collect_sitemap_entries(client, sitemaps) if sitemaps else []
    all_urls = [e["loc"] for e in entries]
    keep, rejected = scope_filter(all_urls, domain, exclude)
    sections = sorted({(parse_locale_section(u)[1] or "(root)") for u in keep})
    return {"domain": domain, "sitemaps": sitemaps, "in_scope": keep,
            "rejected": rejected, "sections": sections,
            "found": len(all_urls), "sitemap_found": bool(sitemaps)}


# ── fetch (robots-aware, throttled) ────────────────────────────────────────
def _robots_allows(client, url: str) -> bool:
    """Cheap robots.txt allow check for our UA (default allow if robots is absent)."""
    try:
        root = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
        r = client.get(urljoin(root, "/robots.txt"))
        if r.status_code >= 400:
            return True
        import urllib.robotparser as rp
        p = rp.RobotFileParser()
        p.parse(r.text.splitlines())
        return p.can_fetch(settings()["crawl"]["user_agent"], url)
    except Exception:  # noqa: BLE001
        return True


# ── the audit engine ───────────────────────────────────────────────────────
def _status(detected_types, vfindings, parse_errors, gfindings) -> str:
    if not detected_types and not parse_errors:
        return "none"
    if parse_errors or any(f["severity"] == "error" for f in vfindings) or \
       any(f["severity"] == "error" for f in gfindings):
        return "errors"
    if vfindings or gfindings:
        return "warnings"
    return "valid"


def _diff_current(current_schema: str | None, live_types: list[str], recommended: list[str]) -> dict | None:
    """Bulk mode: diff the user's SUPPLIED schema against what's live + recommended."""
    if not current_schema:
        return None
    supplied = structured_data(f'<script type="application/ld+json">{current_schema}</script>') \
        if current_schema.strip()[:1] in ("{", "[") else structured_data(current_schema)
    sup_types = set(supplied["types"])
    live = set(live_types)
    return {
        "supplied_types": sorted(sup_types),
        "drift_on_page_not_supplied": sorted(live - sup_types),
        "supplied_not_on_page": sorted(sup_types - live),
        "still_missing_recommended": sorted(set(recommended) - sup_types - live),
    }


def audit_page(conn, source_id: int, url: str, html: str, current_schema: str | None = None) -> dict:
    """Full per-URL audit (reuses the schema_checker analysis library). Pure — no storage."""
    sd = structured_data(html)
    page = page_extract(html, url)
    v = vocab.load()
    vfindings = validate_vocab.validate(sd["nodes"], v)
    gres = google_rules.check(sd["nodes"])
    ia = intent.analyze(html, url, page)
    recs = recommend.recommend(ia["intent"], ia["signals"], sd["types"], vfindings)
    try:
        from ..external.brand import get_brand
        brand = get_brand(conn, source_id) or {}
    except Exception:  # noqa: BLE001
        brand = {}
    gen = generate.generate(url, page, ia["signals"],
                            {"nodes": sd["nodes"], "types": sd["types"],
                             "parse_errors": sd["parse_errors"]}, recs, brand)

    detected = sd["types"]
    recommended_types = [r["type"] for r in recs if r["action"] in ("ADD", "UPGRADE")]
    missing_props = sorted({f["property"] for f in gres["findings"] if f.get("property")})
    status = _status(detected, vfindings, sd["parse_errors"], gres["findings"])
    validity = score.validity_score(vfindings, sd["parse_errors"])
    coverage = score.coverage_score(recs, bool(detected))
    priority = score.priority(validity, coverage, recs)
    return {
        "url": url, "intent": ia["intent"],
        "detected_types": detected or [], "formats": sd["formats"],
        "props": sorted({p for n in sd["nodes"] for p in n["props"]}),
        "block_count": sd["block_count"], "parse_errors": sd["parse_errors"],
        "status": status, "validity": validity, "coverage": coverage,
        "schemaorg_findings": vfindings, "google": gres,
        "recommendations": recs, "recommended_types": recommended_types,
        "missing_props": missing_props, "priority": priority,
        "snippet": gen["pretty"], "snippet_min": gen["minified"],
        "org_warnings": gen["org_warnings"],
        "diff": _diff_current(current_schema, detected, recommended_types),
        "blocks": sd["blocks"],
    }


def _store(conn, source_id, url, result, current_schema):
    """Upsert into page_schema (bound to the project's url row)."""
    locale, section = parse_locale_section(url)
    from ..db import upsert_url
    upsert_url(conn, source_id, url, locale, section, None)
    row = conn.execute("SELECT id FROM urls WHERE source_id=? AND url=?", (source_id, url)).fetchone()
    url_id = row["id"] if row else None
    conn.execute(
        """INSERT INTO page_schema (source_id, url_id, url, format, types_json, props_json,
             status, errors_json, recommended_json, current_schema, priority, created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
           ON CONFLICT(source_id, url_id) DO UPDATE SET
             format=excluded.format, types_json=excluded.types_json, props_json=excluded.props_json,
             status=excluded.status, errors_json=excluded.errors_json,
             recommended_json=excluded.recommended_json, current_schema=excluded.current_schema,
             priority=excluded.priority, created_at=excluded.created_at""",
        (source_id, url_id, url, ",".join(result["formats"]) or "none",
         json.dumps(result["detected_types"]), json.dumps(result["props"]),
         result["status"],
         json.dumps({"schemaorg": result["schemaorg_findings"], "google": result["google"]["findings"]}),
         json.dumps({"types": result["recommended_types"], "missing_props": result["missing_props"],
                     "priority": result["priority"], "snippet": result["snippet"],
                     "diff": result["diff"]}),
         current_schema, result["priority"], now_iso()))


def run_audit(conn, source_id: int, url_rows, on_log=None) -> dict:
    """Audit an explicitly-selected set of (url, current_schema?) rows. Scope-checks AGAIN
    (defence in depth), respects robots + crawl throttle, stores each, returns a per-URL log."""
    domain = scope_domain(conn, source_id)
    exclude = settings()["crawl"].get("exclude_hosts") or []
    c = settings()["crawl"]
    # accept ["url", ...] or [("url", current), ...]
    rows = [(r, None) if isinstance(r, str) else r for r in url_rows]
    keep, rejected = scope_filter([u for u, _ in rows], domain, exclude)
    keep_set = set(keep)
    log, results = [], []
    with httpx.Client(follow_redirects=True, timeout=c["request_timeout_s"],
                      headers={"user-agent": c["user_agent"]}) as client:
        for url, current in rows:
            if url.strip() not in keep_set:
                continue  # rejected by scope (reported separately)
            entry = {"url": url, "ok": False, "message": ""}
            try:
                if not _robots_allows(client, url):
                    entry["message"] = "blocked by robots.txt"
                    log.append(entry); _emit(on_log, entry); continue
                r = client.get(url)
                if r.status_code >= 400 or "html" not in r.headers.get("content-type", ""):
                    entry["message"] = f"HTTP {r.status_code} / non-HTML"
                    log.append(entry); _emit(on_log, entry); continue
                res = audit_page(conn, source_id, url, r.text, current)
                _store(conn, source_id, url, res, current)
                results.append(res)
                entry["ok"] = True
                entry["message"] = f"{res['status']} · {len(res['detected_types'])} type(s) · priority {res['priority']}"
            except Exception as exc:  # noqa: BLE001
                entry["message"] = f"{type(exc).__name__}: {exc}"
            log.append(entry); _emit(on_log, entry)
            time.sleep(c["per_worker_delay_s"])  # reuse crawl throttle
    conn.commit()
    return {"results": results, "log": log,
            "rejected": [{"url": u, "reason": why} for u, why in rejected],
            "audited": len(results), "domain": domain,
            "ruleset_version": google_rules.RULESET["version"]}


def _emit(cb, entry):
    if cb:
        try:
            cb(entry)
        except Exception:  # noqa: BLE001
            pass


def recent(conn, source_id: int, limit: int = 500) -> list[dict]:
    """Stored audit rows for the project (for the dashboard table + export)."""
    rows = conn.execute(
        "SELECT * FROM page_schema WHERE source_id=? ORDER BY id DESC LIMIT ?",
        (source_id, limit)).fetchall()
    out = []
    for r in rows:
        rec = json.loads(r["recommended_json"] or "{}")
        out.append({
            "url": r["url"], "format": r["format"],
            "types": json.loads(r["types_json"] or "[]"),
            "status": r["status"], "priority": r["priority"],
            "recommended_types": rec.get("types", []), "missing_props": rec.get("missing_props", []),
            "snippet": rec.get("snippet", ""), "diff": rec.get("diff"),
            "section": parse_locale_section(r["url"])[1] or "(root)",
        })
    return out

"""Excel (.xlsx) exports — same row data as the CSV exports (no second data path).

Reuses csv_export._issue_rows / external_rows so counts can never diverge from CSV/UI.
Adds Excel niceties: title/scope line, bold+frozen header, autofilter, autosized columns,
real dates, clickable URL hyperlinks, subtle severity/status colour. Unicode-native.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import PROJECT_ROOT, settings
from ..db import CATEGORIES
from . import csv_export as cx

_BRAND_FONT = Font(bold=True, size=16, color="626DF9")  # ExactFact Checker blue
_LOGO_PNG = PROJECT_ROOT / "static" / "img" / "logo-full.png"

_SEV_FILL = {
    "critical": "FDECEE", "high": "FFF2E3", "medium": "FBF3D0", "low": "E7F7EE",
}
_STATUS_FONT = {
    "open": "B42318", "fixed": "0A7D3F", "ignored": "6B7688",
    "false_positive": "6B7688", "unclear": "B25E09",
}
_HEADER_FILL = PatternFill("solid", fgColor="1A2231")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=12, color="1A2231")
_WRAP_COLS = {"url", "external_url", "evidence", "detail", "reason", "expected", "quote", "note",
              "source_url", "context_paragraph", "anchor_text"}
_URL_COLS = {"url", "external_url"}
_DATE_COLS = {"detected_at", "found_at", "created_at"}
_MAXW = {"url": 60, "external_url": 60, "evidence": 70, "expected": 40, "reason": 50,
         "matched_value": 24, "content_type": 12, "author": 22}


def _as_date(v):
    if not v:
        return None
    try:
        return datetime.fromisoformat(str(v).replace("Z", "+00:00")).replace(tzinfo=None)
    except (ValueError, TypeError):
        return v


def _write_sheet(ws, title: str, columns: list[str], rows: list[dict]):
    sev_i = columns.index("severity") + 1 if "severity" in columns else None
    st_i = columns.index("status") + 1 if "status" in columns else None
    # row 1: title/scope; row 2: header; rows 3+: data
    ws.cell(1, 1, title).font = _TITLE_FONT
    for c, name in enumerate(columns, 1):
        cell = ws.cell(2, c, name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    widths = {c: len(name) for c, name in enumerate(columns, 1)}
    for ri, row in enumerate(rows, 3):
        for ci, name in enumerate(columns, 1):
            val = row.get(name)
            cell = ws.cell(ri, ci)
            if name in _DATE_COLS:
                d = _as_date(val)
                cell.value = d
                if isinstance(d, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"
            elif name in _URL_COLS and val:
                cell.value = val
                cell.hyperlink = val
                cell.font = Font(color="2C47D6", underline="single")
            else:
                cell.value = val if val not in (None, "") else None
            if name in _WRAP_COLS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            widths[ci] = max(widths[ci], min(len(str(val or "")), _MAXW.get(name, 30)))
        if sev_i:
            fill = _SEV_FILL.get(str(row.get("severity") or "").lower())
            if fill:
                ws.cell(ri, sev_i).fill = PatternFill("solid", fgColor=fill)
        if st_i:
            col = _STATUS_FONT.get(str(row.get("status") or "").lower())
            if col:
                ws.cell(ri, st_i).font = Font(color=col, bold=True)
    # widths, freeze header, autofilter
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2, 8), 72)
    ws.freeze_panes = "A3"
    last = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A2:{last}{max(2, len(rows) + 2)}"


def _issue_dicts(conn, source_id, category, scope):
    return [{k: r[k] for k in cx.COLS} for r in cx._issue_rows(conn, source_id, category, scope)]


def _title(conn, source_id, scope, label):
    src = conn.execute("SELECT name, location FROM sources WHERE id=?", (source_id,)).fetchone()
    proj = (src["name"] or src["location"]) if src else source_id
    return f"{proj} — {label} — scope={scope} — generated {datetime.now():%Y-%m-%d %H:%M}"


def build_issue_xlsx(conn, source_id: int, category: str | None = None, scope: str = "open") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (category or "findings")[:31]
    _write_sheet(ws, _title(conn, source_id, scope, category or "All findings"),
                 cx.COLS, _issue_dicts(conn, source_id, category, scope))
    return _save(wb)


def build_url_summary_xlsx(conn, source_id: int, scope: str = "open") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "By URL"
    _write_sheet(ws, _title(conn, source_id, scope, "One row per URL (aggregated)"),
                 cx.URL_COLS, cx.url_summary_rows(conn, source_id, scope))
    return _save(wb)


def build_external_xlsx(conn, source_id: int, scope: str = "open") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "External"
    _write_sheet(ws, _title(conn, source_id, scope, "External findings"),
                 cx.EXT_COLS, cx.external_rows(conn, source_id, scope))
    return _save(wb)


def build_rows_xlsx(rows, ext_rows) -> bytes:
    """Filtered Results view (already-fetched dicts) — mirrors build_rows_csv."""
    cols = ["scope", *cx.COLS]
    data = []
    for r in rows:
        data.append({"scope": "internal", "url": r["url"], "locale": r.get("locale") or "",
                     "content_type": r.get("content_type") or "", "author": r.get("author") or "",
                     "product": r.get("product_name") or "", "category": r["category"],
                     "severity": r["severity"], "status": r["status"], "fact_rule": r.get("title") or "",
                     "evidence": r.get("evidence") or "", "matched_value": r.get("matched_value") or "",
                     "expected": r.get("expected") or "",
                     "detection_method": r.get("detection_method") or "", "detected_at": r.get("last_checked_at") or ""})
    for r in ext_rows:
        data.append({"scope": "external", "url": r["external_url"], "locale": "", "content_type": "",
                     "author": "", "product": "",
                     "category": r.get("finding_type") or "", "severity": r["severity"], "status": r["status"],
                     "fact_rule": "", "evidence": r.get("snippet") or "", "matched_value": "",
                     "expected": r.get("expected") or "",
                     "detection_method": "", "detected_at": ""})
    wb = Workbook(); ws = wb.active; ws.title = "Findings"
    _write_sheet(ws, f"Filtered findings — generated {datetime.now():%Y-%m-%d %H:%M}", cols, data)
    return _save(wb)


def build_external_items_xlsx(conn, source_id: int, aliases=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlinks & Mentions"
    rows = cx.external_item_rows(conn, source_id, aliases)
    _write_sheet(ws, _title(conn, source_id, "all", "Backlinks & Mentions (context report)"),
                 cx.EXTITEM_COLS, rows)
    return _save(wb)


def build_general_facts_xlsx(conn, source_id: int, scope: str = "open") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "General Facts"
    rows = [{k: r[k] for k in cx.GENFACT_COLS} for r in cx.general_fact_rows(conn, source_id, scope)]
    _write_sheet(ws, _title(conn, source_id, scope, "General Facts (brand info to review)"),
                 cx.GENFACT_COLS, rows)
    return _save(wb)


def build_full_report_xlsx(conn, source_id: int, scope: str = "open") -> bytes:
    """Multi-sheet workbook: Summary tab first, then one tab per fact category."""
    from ..db import source_domain
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    # gather per-category counts
    cat_rows = []
    for cat in CATEGORIES:
        openc = conn.execute("SELECT COUNT(*) c FROM issues WHERE source_id=? AND category=? "
                             "AND status='open' AND deleted_at IS NULL", (source_id, cat)).fetchone()["c"]
        allc = conn.execute("SELECT COUNT(*) c FROM issues WHERE source_id=? AND category=? "
                            "AND deleted_at IS NULL", (source_id, cat)).fetchone()["c"]
        if allc:
            cat_rows.append({"category": cat, "open": openc, "all": allc})
    sev = {s: conn.execute("SELECT COUNT(*) c FROM issues WHERE source_id=? AND severity=? "
                           "AND status='open' AND deleted_at IS NULL", (source_id, s)).fetchone()["c"]
           for s in ("critical", "high", "medium", "low")}
    crawled = conn.execute("SELECT COUNT(DISTINCT c.url_id) c FROM crawl_results c JOIN urls u ON u.id=c.url_id "
                           "WHERE u.source_id=?", (source_id,)).fetchone()["c"]
    total = conn.execute("SELECT COUNT(*) c FROM urls WHERE source_id=? AND in_source=1", (source_id,)).fetchone()["c"]

    # Summary sheet — branded header (ExactFact Checker by SDS Manager)
    summary.cell(1, 1, "ExactFact Checker").font = _BRAND_FONT
    summary.cell(2, 1, "by SDS Manager · Full fact-check report").font = Font(bold=True, color="6B7688")
    summary.cell(3, 1, _title(conn, source_id, scope, "Full fact-check report")).font = Font(color="6B7688")
    try:  # float the full logo top-right (PNG fallback; never break the export)
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(str(_LOGO_PNG))
        img.width, img.height = 150, 50
        summary.add_image(img, "D1")
    except Exception:  # noqa: BLE001
        pass
    r = 5
    summary.cell(r, 1, "Coverage").font = Font(bold=True); r += 1
    summary.cell(r, 1, "URLs in project"); summary.cell(r, 2, total); r += 1
    summary.cell(r, 1, "URLs crawled"); summary.cell(r, 2, crawled); r += 2
    summary.cell(r, 1, "Open findings by severity").font = Font(bold=True); r += 1
    for s, n in sev.items():
        summary.cell(r, 1, s); summary.cell(r, 2, n); r += 1
    r += 1
    hdr = summary.cell(r, 1, "Category"); hdr.font = _HEADER_FONT; hdr.fill = _HEADER_FILL
    for j, t in enumerate(["Open", "All statuses"], 2):
        cc = summary.cell(r, j, t); cc.font = _HEADER_FONT; cc.fill = _HEADER_FILL
    r += 1
    for cr in cat_rows:
        summary.cell(r, 1, cr["category"]); summary.cell(r, 2, cr["open"]); summary.cell(r, 3, cr["all"]); r += 1
    for ci, w in {1: 26, 2: 12, 3: 14}.items():
        summary.column_dimensions[get_column_letter(ci)].width = w
    summary.freeze_panes = "A5"

    # one sheet per category with data
    used = {"Summary"}
    for cr in cat_rows:
        cat = cr["category"]
        name = cat[:31] or "cat"
        while name in used:
            name = name[:29] + "_2"
        used.add(name)
        ws = wb.create_sheet(name)
        _write_sheet(ws, _title(conn, source_id, scope, cat), cx.COLS,
                     _issue_dicts(conn, source_id, cat, scope))
    return _save(wb)


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
